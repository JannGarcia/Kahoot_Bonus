import os
import openpyxl

# Constants
REPORTS_DIR = "reports"
SUMMARY_FILE = "bonus_points_summary.xlsx"
MAX_STUDENTS_PER_REPORT = 20
POINTS_FOR_BEING_IN_TOP = 5
MAX_POINTS_PER_STUDENT = 50


def get_rows_from_report(report_file):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(report_file)
    sheet = wb["Final Scores"]

    # Get all rows starting from the third row (to skip headers)
    return sheet.iter_rows(min_row=4, values_only=True)


"""
report_rows = [
    0. Rank	
    1. Player	<-- we only care about this column
    2. Total Score (points)	
    3. Correct Answers	
    4. Incorrect Answers
]
"""


def calculate_bonus_points(report_rows, points_dict):
    students_added = 0

    for row in report_rows:
        player = row[1].lower()

        # Ensure username has the correct format of "firstname.lastname"
        if "." not in player:
            continue

        # Stop adding students if the maximum number of students have been added for this report
        if students_added >= MAX_STUDENTS_PER_REPORT:
            break

        student_points = points_dict.get(player, 0)

        # Check if the player is in the top k and has less than m points
        if student_points < MAX_POINTS_PER_STUDENT:
            # Calculate bonus points
            bonus_points = min(
                MAX_POINTS_PER_STUDENT - student_points, POINTS_FOR_BEING_IN_TOP
            )
            points_dict[player] = student_points + bonus_points
            students_added += 1


def save_summary_to_excel(points_dict):
    # Sort the dictionary by keys (usernames)
    sorted_points = sorted(points_dict.items())

    # Create a new workbook to save the summary
    result_wb = openpyxl.Workbook()
    result_sheet = result_wb.active

    # Write headers
    result_sheet["A1"] = "Name"
    result_sheet["B1"] = "Bonus Points Earned"

    # Write data
    for i, (player, points) in enumerate(sorted_points, start=2):
        result_sheet[f"A{i}"] = player
        result_sheet[f"B{i}"] = points

    # Save the summary to a new file
    result_wb.save(SUMMARY_FILE)
    print(f"Summary saved to {SUMMARY_FILE}")


def main():

    # Directory containing reports
    reports_dir = REPORTS_DIR

    # Dictionary to store points for each student
    points_dict = {}

    # Process each report file in the directory
    for filename in os.listdir(reports_dir):
        if filename.endswith(".xlsx"):
            report_file = os.path.join(reports_dir, filename)
            print(f"Processing {report_file}...")
            calculate_bonus_points(get_rows_from_report(report_file), points_dict)
            print()

    # Save the summary to a single Excel file
    save_summary_to_excel(points_dict)


if __name__ == "__main__":
    main()
